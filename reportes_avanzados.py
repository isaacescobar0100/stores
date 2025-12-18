#!/usr/bin/env python3
"""
Módulo de Reportes Avanzados para Superadmin
Incluye: Excel export, reportes por hora, por método de pago, comparativos
"""
from flask import Blueprint, request, Response, jsonify, render_template
from datetime import datetime, timedelta
from decimal import Decimal
import io
import csv

reportes_bp = Blueprint('reportes', __name__, url_prefix='/superadmin/reportes')


# ============ VISTA PRINCIPAL ============

@reportes_bp.route('/')
def reportes_index():
    """Vista principal de reportes avanzados"""
    from flask import session, redirect, url_for
    if not session.get('superadmin'):
        return redirect(url_for('superadmin.login'))

    db = get_connection()
    cursor = db.cursor()

    fecha_hasta = request.args.get('hasta', datetime.now().strftime('%Y-%m-%d'))
    fecha_desde = request.args.get('desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    tienda_id = request.args.get('tienda')

    # Obtener lista de tiendas
    cursor.execute('SELECT id, nombre FROM tiendas WHERE activo = 1 ORDER BY nombre')
    tiendas = cursor.fetchall()

    cursor.close()
    db.close()

    return render_template('superadmin/reportes_avanzados.html',
                         tiendas=tiendas,
                         fecha_desde=fecha_desde,
                         fecha_hasta=fecha_hasta,
                         tienda_seleccionada=int(tienda_id) if tienda_id else None)


def get_connection():
    """Obtener conexión a la base de datos"""
    import pymysql
    import os
    return pymysql.connect(
        host=os.environ.get('DB_HOST', 'mysql'),
        user=os.environ.get('DB_USER', 'restaurante'),
        password=os.environ.get('DB_PASSWORD', 'restaurante_pass_2024'),
        database=os.environ.get('DB_NAME', 'restaurantes'),
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True
    )


def requiere_superadmin(f):
    """Decorator para verificar acceso superadmin"""
    from functools import wraps
    from flask import session, redirect, url_for
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('superadmin'):
            return redirect(url_for('superadmin.login'))
        return f(*args, **kwargs)
    return decorated


# ============ EXPORTAR A EXCEL ============

@reportes_bp.route('/exportar/excel')
@requiere_superadmin
def exportar_excel():
    """Exportar reporte completo a Excel con múltiples hojas"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_connection()
    cursor = db.cursor()

    fecha_desde = request.args.get('desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('hasta', datetime.now().strftime('%Y-%m-%d'))
    tienda_id = request.args.get('tienda')

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== HOJA 1: RESUMEN GENERAL =====
    ws1 = wb.active
    ws1.title = "Resumen"

    # Query stats generales
    query_stats = '''
        SELECT
            COUNT(DISTINCT p.id) as total_pedidos,
            COALESCE(SUM(p.total), 0) as total_ventas,
            COUNT(DISTINCT p.tienda_id) as tiendas_con_ventas,
            COUNT(DISTINCT p.cliente_id) as clientes_unicos
        FROM pedidos p
        WHERE DATE(p.fecha_pedido) BETWEEN %s AND %s
          AND p.estado != 'cancelado'
    '''
    params = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_stats += ' AND p.tienda_id = %s'
        params.append(tienda_id)

    cursor.execute(query_stats, params)
    stats = cursor.fetchone()

    # Escribir resumen
    ws1['A1'] = 'REPORTE DE VENTAS'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = f'Período: {fecha_desde} a {fecha_hasta}'
    ws1['A4'] = 'Métrica'
    ws1['B4'] = 'Valor'
    ws1['A4'].font = header_font
    ws1['B4'].font = header_font
    ws1['A4'].fill = header_fill
    ws1['B4'].fill = header_fill

    ws1['A5'] = 'Total Ventas'
    ws1['B5'] = float(stats['total_ventas'])
    ws1['B5'].number_format = '"$"#,##0.00'
    ws1['A6'] = 'Total Pedidos'
    ws1['B6'] = stats['total_pedidos']
    ws1['A7'] = 'Ticket Promedio'
    ws1['B7'] = float(stats['total_ventas'] / stats['total_pedidos']) if stats['total_pedidos'] > 0 else 0
    ws1['B7'].number_format = '"$"#,##0.00'
    ws1['A8'] = 'Tiendas con Ventas'
    ws1['B8'] = stats['tiendas_con_ventas']
    ws1['A9'] = 'Clientes Únicos'
    ws1['B9'] = stats['clientes_unicos']

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15

    # ===== HOJA 2: VENTAS POR TIENDA =====
    ws2 = wb.create_sheet("Ventas por Tienda")

    query_tiendas = '''
        SELECT t.nombre, t.subdominio,
               COUNT(p.id) as pedidos,
               COALESCE(SUM(p.total), 0) as ventas
        FROM tiendas t
        LEFT JOIN pedidos p ON t.id = p.tienda_id
            AND DATE(p.fecha_pedido) BETWEEN %s AND %s
            AND p.estado != 'cancelado'
        WHERE t.activo = 1
    '''
    params_tiendas = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_tiendas += ' AND t.id = %s'
        params_tiendas.append(tienda_id)
    query_tiendas += ' GROUP BY t.id ORDER BY ventas DESC'

    cursor.execute(query_tiendas, params_tiendas)
    ventas_tiendas = cursor.fetchall()

    headers = ['Tienda', 'Subdominio', 'Pedidos', 'Ventas', '% del Total']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    total_general = sum(float(t['ventas']) for t in ventas_tiendas)
    for row, tienda in enumerate(ventas_tiendas, 2):
        ws2.cell(row=row, column=1, value=tienda['nombre']).border = border
        ws2.cell(row=row, column=2, value=tienda['subdominio']).border = border
        ws2.cell(row=row, column=3, value=tienda['pedidos']).border = border
        cell_ventas = ws2.cell(row=row, column=4, value=float(tienda['ventas']))
        cell_ventas.number_format = '"$"#,##0.00'
        cell_ventas.border = border
        porcentaje = (float(tienda['ventas']) / total_general * 100) if total_general > 0 else 0
        cell_pct = ws2.cell(row=row, column=5, value=porcentaje)
        cell_pct.number_format = '0.0"%"'
        cell_pct.border = border

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # ===== HOJA 3: VENTAS POR HORA =====
    ws3 = wb.create_sheet("Ventas por Hora")

    query_hora = '''
        SELECT HOUR(fecha_pedido) as hora,
               COUNT(*) as pedidos,
               SUM(total) as ventas
        FROM pedidos
        WHERE DATE(fecha_pedido) BETWEEN %s AND %s
          AND estado != 'cancelado'
    '''
    params_hora = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_hora += ' AND tienda_id = %s'
        params_hora.append(tienda_id)
    query_hora += ' GROUP BY HOUR(fecha_pedido) ORDER BY hora'

    cursor.execute(query_hora, params_hora)
    ventas_hora = cursor.fetchall()

    headers_hora = ['Hora', 'Pedidos', 'Ventas', '% Pedidos']
    for col, header in enumerate(headers_hora, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    total_pedidos_hora = sum(h['pedidos'] for h in ventas_hora)
    for row, hora in enumerate(ventas_hora, 2):
        ws3.cell(row=row, column=1, value=f"{hora['hora']:02d}:00 - {hora['hora']:02d}:59")
        ws3.cell(row=row, column=2, value=hora['pedidos'])
        cell_v = ws3.cell(row=row, column=3, value=float(hora['ventas']))
        cell_v.number_format = '"$"#,##0.00'
        pct = (hora['pedidos'] / total_pedidos_hora * 100) if total_pedidos_hora > 0 else 0
        cell_p = ws3.cell(row=row, column=4, value=pct)
        cell_p.number_format = '0.0"%"'

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # ===== HOJA 4: VENTAS POR TIPO (Domicilio/Recoger) =====
    ws4 = wb.create_sheet("Ventas por Tipo")

    query_tipo = '''
        SELECT tipo,
               COUNT(*) as pedidos,
               SUM(total) as ventas
        FROM pedidos
        WHERE DATE(fecha_pedido) BETWEEN %s AND %s
          AND estado != 'cancelado'
    '''
    params_tipo = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_tipo += ' AND tienda_id = %s'
        params_tipo.append(tienda_id)
    query_tipo += ' GROUP BY tipo'

    cursor.execute(query_tipo, params_tipo)
    ventas_tipo = cursor.fetchall()

    headers_tipo = ['Tipo', 'Pedidos', 'Ventas', '% Ventas']
    for col, header in enumerate(headers_tipo, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    total_ventas_tipo = sum(float(t['ventas']) for t in ventas_tipo)
    for row, tipo in enumerate(ventas_tipo, 2):
        nombre_tipo = 'Domicilio' if tipo['tipo'] == 'domicilio' else 'Recoger en Tienda'
        ws4.cell(row=row, column=1, value=nombre_tipo)
        ws4.cell(row=row, column=2, value=tipo['pedidos'])
        cell_v = ws4.cell(row=row, column=3, value=float(tipo['ventas']))
        cell_v.number_format = '"$"#,##0.00'
        pct = (float(tipo['ventas']) / total_ventas_tipo * 100) if total_ventas_tipo > 0 else 0
        cell_p = ws4.cell(row=row, column=4, value=pct)
        cell_p.number_format = '0.0"%"'

    # ===== HOJA 5: PRODUCTOS MÁS VENDIDOS =====
    ws5 = wb.create_sheet("Top Productos")

    query_productos = '''
        SELECT pr.nombre, t.nombre as tienda,
               SUM(dp.cantidad) as cantidad,
               SUM(dp.cantidad * dp.precio_unitario) as total
        FROM detalle_pedidos dp
        JOIN pedidos p ON dp.pedido_id = p.id
        JOIN productos pr ON dp.producto_id = pr.id
        JOIN tiendas t ON p.tienda_id = t.id
        WHERE DATE(p.fecha_pedido) BETWEEN %s AND %s
          AND p.estado != 'cancelado'
    '''
    params_prod = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_productos += ' AND p.tienda_id = %s'
        params_prod.append(tienda_id)
    query_productos += ' GROUP BY pr.id ORDER BY cantidad DESC LIMIT 50'

    cursor.execute(query_productos, params_prod)
    productos = cursor.fetchall()

    headers_prod = ['Producto', 'Tienda', 'Cantidad', 'Total Vendido']
    for col, header in enumerate(headers_prod, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, prod in enumerate(productos, 2):
        ws5.cell(row=row, column=1, value=prod['nombre'])
        ws5.cell(row=row, column=2, value=prod['tienda'])
        ws5.cell(row=row, column=3, value=int(prod['cantidad']))
        cell_t = ws5.cell(row=row, column=4, value=float(prod['total']))
        cell_t.number_format = '"$"#,##0.00'

    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 20

    # ===== HOJA 6: VENTAS DIARIAS =====
    ws6 = wb.create_sheet("Ventas Diarias")

    query_diario = '''
        SELECT DATE(fecha_pedido) as fecha,
               COUNT(*) as pedidos,
               SUM(total) as ventas
        FROM pedidos
        WHERE DATE(fecha_pedido) BETWEEN %s AND %s
          AND estado != 'cancelado'
    '''
    params_diario = [fecha_desde, fecha_hasta]
    if tienda_id:
        query_diario += ' AND tienda_id = %s'
        params_diario.append(tienda_id)
    query_diario += ' GROUP BY DATE(fecha_pedido) ORDER BY fecha'

    cursor.execute(query_diario, params_diario)
    ventas_diarias = cursor.fetchall()

    headers_diario = ['Fecha', 'Día', 'Pedidos', 'Ventas']
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    for col, header in enumerate(headers_diario, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, dia in enumerate(ventas_diarias, 2):
        fecha_obj = dia['fecha']
        ws6.cell(row=row, column=1, value=fecha_obj.strftime('%Y-%m-%d') if hasattr(fecha_obj, 'strftime') else str(fecha_obj))
        ws6.cell(row=row, column=2, value=dias_semana[fecha_obj.weekday()] if hasattr(fecha_obj, 'weekday') else '')
        ws6.cell(row=row, column=3, value=dia['pedidos'])
        cell_v = ws6.cell(row=row, column=4, value=float(dia['ventas']))
        cell_v.number_format = '"$"#,##0.00'

    # Generar archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    cursor.close()
    db.close()

    filename = f'reporte_ventas_{fecha_desde}_{fecha_hasta}.xlsx'
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ============ API: VENTAS POR HORA =====

@reportes_bp.route('/api/ventas-por-hora')
@requiere_superadmin
def api_ventas_por_hora():
    """Obtener ventas agrupadas por hora del día"""
    db = get_connection()
    cursor = db.cursor()

    fecha_desde = request.args.get('desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('hasta', datetime.now().strftime('%Y-%m-%d'))
    tienda_id = request.args.get('tienda')

    query = '''
        SELECT HOUR(fecha_pedido) as hora,
               COUNT(*) as pedidos,
               SUM(total) as ventas
        FROM pedidos
        WHERE DATE(fecha_pedido) BETWEEN %s AND %s
          AND estado != 'cancelado'
    '''
    params = [fecha_desde, fecha_hasta]
    if tienda_id:
        query += ' AND tienda_id = %s'
        params.append(tienda_id)
    query += ' GROUP BY HOUR(fecha_pedido) ORDER BY hora'

    cursor.execute(query, params)
    resultados = cursor.fetchall()

    # Llenar horas vacías
    horas_completas = []
    horas_dict = {r['hora']: r for r in resultados}
    for h in range(24):
        if h in horas_dict:
            horas_completas.append({
                'hora': h,
                'label': f'{h:02d}:00',
                'pedidos': horas_dict[h]['pedidos'],
                'ventas': float(horas_dict[h]['ventas'])
            })
        else:
            horas_completas.append({
                'hora': h,
                'label': f'{h:02d}:00',
                'pedidos': 0,
                'ventas': 0
            })

    cursor.close()
    db.close()

    return jsonify(horas_completas)


# ============ API: COMPARATIVO MENSUAL =====

@reportes_bp.route('/api/comparativo-mensual')
@requiere_superadmin
def api_comparativo_mensual():
    """Comparar mes actual vs mes anterior"""
    db = get_connection()
    cursor = db.cursor()

    tienda_id = request.args.get('tienda')

    hoy = datetime.now()
    inicio_mes_actual = hoy.replace(day=1)
    fin_mes_anterior = inicio_mes_actual - timedelta(days=1)
    inicio_mes_anterior = fin_mes_anterior.replace(day=1)

    def get_stats_periodo(inicio, fin):
        query = '''
            SELECT
                COUNT(*) as pedidos,
                COALESCE(SUM(total), 0) as ventas,
                COUNT(DISTINCT cliente_id) as clientes
            FROM pedidos
            WHERE DATE(fecha_pedido) BETWEEN %s AND %s
              AND estado != 'cancelado'
        '''
        params = [inicio.strftime('%Y-%m-%d'), fin.strftime('%Y-%m-%d')]
        if tienda_id:
            query += ' AND tienda_id = %s'
            params.append(tienda_id)
        cursor.execute(query, params)
        return cursor.fetchone()

    stats_actual = get_stats_periodo(inicio_mes_actual, hoy)
    stats_anterior = get_stats_periodo(inicio_mes_anterior, fin_mes_anterior)

    def calcular_cambio(actual, anterior):
        if anterior == 0:
            return 100 if actual > 0 else 0
        return round((actual - anterior) / anterior * 100, 1)

    resultado = {
        'mes_actual': {
            'nombre': hoy.strftime('%B %Y'),
            'pedidos': stats_actual['pedidos'],
            'ventas': float(stats_actual['ventas']),
            'clientes': stats_actual['clientes'],
            'ticket_promedio': float(stats_actual['ventas'] / stats_actual['pedidos']) if stats_actual['pedidos'] > 0 else 0
        },
        'mes_anterior': {
            'nombre': fin_mes_anterior.strftime('%B %Y'),
            'pedidos': stats_anterior['pedidos'],
            'ventas': float(stats_anterior['ventas']),
            'clientes': stats_anterior['clientes'],
            'ticket_promedio': float(stats_anterior['ventas'] / stats_anterior['pedidos']) if stats_anterior['pedidos'] > 0 else 0
        },
        'cambio': {
            'pedidos': calcular_cambio(stats_actual['pedidos'], stats_anterior['pedidos']),
            'ventas': calcular_cambio(float(stats_actual['ventas']), float(stats_anterior['ventas'])),
            'clientes': calcular_cambio(stats_actual['clientes'], stats_anterior['clientes'])
        }
    }

    cursor.close()
    db.close()

    return jsonify(resultado)


# ============ API: VENTAS POR DÍA DE SEMANA =====

@reportes_bp.route('/api/ventas-por-dia-semana')
@requiere_superadmin
def api_ventas_por_dia_semana():
    """Ventas agrupadas por día de la semana"""
    db = get_connection()
    cursor = db.cursor()

    fecha_desde = request.args.get('desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('hasta', datetime.now().strftime('%Y-%m-%d'))
    tienda_id = request.args.get('tienda')

    query = '''
        SELECT DAYOFWEEK(fecha_pedido) as dia,
               COUNT(*) as pedidos,
               SUM(total) as ventas
        FROM pedidos
        WHERE DATE(fecha_pedido) BETWEEN %s AND %s
          AND estado != 'cancelado'
    '''
    params = [fecha_desde, fecha_hasta]
    if tienda_id:
        query += ' AND tienda_id = %s'
        params.append(tienda_id)
    query += ' GROUP BY DAYOFWEEK(fecha_pedido) ORDER BY dia'

    cursor.execute(query, params)
    resultados = cursor.fetchall()

    dias_nombres = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    dias_dict = {r['dia']: r for r in resultados}

    dias_completos = []
    for d in range(1, 8):  # 1=Domingo, 7=Sábado en MySQL
        if d in dias_dict:
            dias_completos.append({
                'dia': d,
                'nombre': dias_nombres[d-1],
                'pedidos': dias_dict[d]['pedidos'],
                'ventas': float(dias_dict[d]['ventas'])
            })
        else:
            dias_completos.append({
                'dia': d,
                'nombre': dias_nombres[d-1],
                'pedidos': 0,
                'ventas': 0
            })

    cursor.close()
    db.close()

    return jsonify(dias_completos)


# ============ API: TOP CLIENTES =====

@reportes_bp.route('/api/top-clientes')
@requiere_superadmin
def api_top_clientes():
    """Top clientes por ventas"""
    db = get_connection()
    cursor = db.cursor()

    fecha_desde = request.args.get('desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('hasta', datetime.now().strftime('%Y-%m-%d'))
    tienda_id = request.args.get('tienda')
    limite = request.args.get('limite', 20, type=int)

    query = '''
        SELECT c.nombre, c.telefono, c.email,
               COUNT(p.id) as total_pedidos,
               SUM(p.total) as total_compras,
               MAX(p.fecha_pedido) as ultima_compra
        FROM clientes c
        JOIN pedidos p ON c.id = p.cliente_id
        WHERE DATE(p.fecha_pedido) BETWEEN %s AND %s
          AND p.estado != 'cancelado'
    '''
    params = [fecha_desde, fecha_hasta]
    if tienda_id:
        query += ' AND p.tienda_id = %s'
        params.append(tienda_id)
    query += f' GROUP BY c.id ORDER BY total_compras DESC LIMIT {limite}'

    cursor.execute(query, params)
    clientes = cursor.fetchall()

    resultado = []
    for c in clientes:
        resultado.append({
            'nombre': c['nombre'],
            'telefono': c['telefono'],
            'email': c['email'],
            'pedidos': c['total_pedidos'],
            'total': float(c['total_compras']),
            'ultima_compra': c['ultima_compra'].strftime('%Y-%m-%d %H:%M') if c['ultima_compra'] else None
        })

    cursor.close()
    db.close()

    return jsonify(resultado)
